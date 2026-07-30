import streamlit as st
import openpyxl
from io import BytesIO
import pandas as pd
import os
import datetime
import zipfile
import base64

# --- Konfigurasi ---
TEMPLATE_F13 = "F 13 - Kuesioner Pelanggan.xlsx"
TEMPLATE_F15 = "F15_Template.xlsx" 
NAMA_SHEET_F13 = "rev 3"
NAMA_SHEET_F15 = "master rev 2 (2)"
LOG_FILE = "log_kuesioner_baru.csv" 
FOLDER_HASIL = "hasil_kuesioner"
LOGO_FILE = "logo tkr.jpg" 
DATA_F15 = "data_evaluasi.csv"

# Membuat folder dan file log
if not os.path.exists(FOLDER_HASIL):
    os.makedirs(FOLDER_HASIL)
if not os.path.exists(LOG_FILE):
    pd.DataFrame(columns=["Nama / Instansi", "Tujuan Uji"]).to_csv(LOG_FILE, index=False)
if not os.path.exists(DATA_F15):
    kolom = ["h_1", "h_2", "h_3", "h_4", "h_5", "h_6", "h_7", "h_8", "h_9", "h_10", "h_11", "h_12", "h_13", "h_14",
             "k_1", "k_2", "k_3", "k_4", "k_5", "k_6", "k_7", "k_8", "k_9", "k_10", "k_11", "k_12", "k_13", "k_14"]
    pd.DataFrame(columns=kolom).to_csv(DATA_F15, index=False)

st.set_page_config(page_title="Kuesioner Lab TKR", layout="centered")

def buat_watermark():
    if os.path.exists(LOGO_FILE):
        with open(LOGO_FILE, "rb") as f:
            encoded_string = base64.b64encode(f.read()).decode()
        st.markdown(
            f"""<style>.stApp::before {{content: "";background-image: url(data:image/jpeg;base64,{encoded_string});background-size: 350px;background-position: center;background-repeat: no-repeat;background-attachment: fixed;position: fixed;top: 50%;left: 50%;transform: translate(-50%, -50%);width: 100vw;height: 100vh;opacity: 0.07; z-index: -1;pointer-events: none; }}</style>""",
            unsafe_allow_html=True
        )

buat_watermark()

if os.path.exists(LOGO_FILE):
    st.sidebar.image(LOGO_FILE, use_container_width=True)

st.sidebar.title("Navigasi")
menu = st.sidebar.radio("Pilih Halaman:", ["Form Kuesioner", "Panel Admin"])

# ==========================================
# MANAJEMEN MEMORI CERDAS (ANTI HILANG)
# ==========================================
if "step" not in st.session_state:
    st.session_state.step = 1

aspek_list = [(27, 'x1', 'Kemudahan mencapai lokasi'), (28, 'x2', 'Kejelasan Papan Nama'), (29, 'x3', 'Kenyamanan ruang'), (30, 'x4', 'Sarana Parkir'), (32, 'x5', 'Keramahan Petugas'), (33, 'x6', 'Layanan telepon/fax'), (35, 'x7', 'Kepercayaan hasil uji'), (36, 'x8', 'Peralatan lengkap & modern'), (37, 'x9', 'Akreditasi KAN'), (38, 'x10', 'Kemampuan petugas'), (40, 'x11', 'Parameter standar Permenkes'), (41, 'x12', 'Tampilan LHP'), (42, 'x13', 'Prosedur Pengujian'), (43, 'x14', 'Ketepatan Waktu')]

# Inisialisasi Memori Bagian A (Default 1)
for baris, _, _ in aspek_list:
    if f"h_{baris}" not in st.session_state: st.session_state[f"h_{baris}"] = 1 
    if f"l_{baris}" not in st.session_state: st.session_state[f"l_{baris}"] = 1 

# Inisialisasi Memori Bagian B
if "q1" not in st.session_state: st.session_state.q1 = "Usaha"
if "q2" not in st.session_state: st.session_state.q2 = "Laboratorium PDAM TKR"
if "q2_alasan" not in st.session_state: st.session_state.q2_alasan = ""
if "q3" not in st.session_state: st.session_state.q3 = "Ya"
if "q3_nama" not in st.session_state: st.session_state.q3_nama = ""
if "q3_telp" not in st.session_state: st.session_state.q3_telp = ""
if "q3_alamat" not in st.session_state: st.session_state.q3_alamat = ""
if "q4" not in st.session_state: st.session_state.q4 = "Cukup"
if "q4_lainnya" not in st.session_state: st.session_state.q4_lainnya = ""
if "q5_saran" not in st.session_state: st.session_state.q5_saran = ""
param_kurang = {'Amoniak': 70, 'Aluminium': 71, 'Seng': 72, 'Tembaga': 73, 'Detergent': 74, 'Kadmium': 75, 'Chromium Valensi 6': 76, 'Sianida': 77, 'Flourida': 78, 'Phospat': 79}
for param, baris in param_kurang.items():
    if f"param_{baris}" not in st.session_state: st.session_state[f"param_{baris}"] = False

# Fungsi Penyelamat Data Sebelum Pindah Halaman
def simpan_a():
    for baris, _, _ in aspek_list:
        if f"wid_h_{baris}" in st.session_state: st.session_state[f"h_{baris}"] = st.session_state[f"wid_h_{baris}"]
        if f"wid_l_{baris}" in st.session_state: st.session_state[f"l_{baris}"] = st.session_state[f"wid_l_{baris}"]

def simpan_b():
    if "w_q1" in st.session_state: st.session_state.q1 = st.session_state.w_q1
    if "w_q2" in st.session_state: st.session_state.q2 = st.session_state.w_q2
    if "w_q2_alasan" in st.session_state: st.session_state.q2_alasan = st.session_state.w_q2_alasan
    if "w_q3" in st.session_state: st.session_state.q3 = st.session_state.w_q3
    if "w_q3_nama" in st.session_state: st.session_state.q3_nama = st.session_state.w_q3_nama
    if "w_q3_telp" in st.session_state: st.session_state.q3_telp = st.session_state.w_q3_telp
    if "w_q3_alamat" in st.session_state: st.session_state.q3_alamat = st.session_state.w_q3_alamat
    if "w_q4" in st.session_state: st.session_state.q4 = st.session_state.w_q4
    if "w_q4_lainnya" in st.session_state: st.session_state.q4_lainnya = st.session_state.w_q4_lainnya
    if "w_q5_saran" in st.session_state: st.session_state.q5_saran = st.session_state.w_q5_saran
    for param, baris in param_kurang.items():
        if f"w_p_{baris}" in st.session_state: st.session_state[f"param_{baris}"] = st.session_state[f"w_p_{baris}"]

def go_next(): 
    simpan_a()
    st.session_state.step = 2

def go_back(): 
    simpan_b()
    st.session_state.step = 1

def reset_kuesioner():
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.session_state.step = 1

# ==========================================
# HALAMAN 1: FORM KUESIONER 
# ==========================================
if menu == "Form Kuesioner":
    st.title("Kuesioner Kepuasan Pelanggan")
    st.subheader("Laboratorium Perumdam Tirta Kerta Raharja")

    # --- TAHAP 1: BAGIAN A ---
    if st.session_state.step == 1:
        st.write("### Bagian A: Penilaian Kinerja Laboratorium")
        st.info("Geser indikator berikut untuk memberikan nilai dari 1 (Tidak Baik/Penting) hingga 4 (Sangat Baik/Penting).")
        
        for idx, (baris, kode, teks) in enumerate(aspek_list):
            st.write(f"**{kode}. {teks}**")
            col1, col2 = st.columns(2)
            with col1: st.slider("Harapan Anda", 1, 4, value=st.session_state[f"h_{baris}"], key=f"wid_h_{baris}")
            with col2: st.slider("Pelayanan Dirasakan", 1, 4, value=st.session_state[f"l_{baris}"], key=f"wid_l_{baris}")
            st.write("---")
        
        st.button("Berikutnya ➡️", type="primary", on_click=go_next)

    # --- TAHAP 2: BAGIAN B ---
    elif st.session_state.step == 2:
        st.write("### Bagian B: Profil & Kebutuhan Pelanggan")
        st.radio("1. Kepentingan pengujian:", ["Usaha", "Non Usaha"], index=["Usaha", "Non Usaha"].index(st.session_state.q1), key="w_q1")
        st.radio("2. Pilihan laboratorium:", ["Laboratorium PDAM TKR", "Laboratorium Lain"], index=["Laboratorium PDAM TKR", "Laboratorium Lain"].index(st.session_state.q2), key="w_q2")
        st.text_input("Alasan memilih:", value=st.session_state.q2_alasan, key="w_q2_alasan")
        
        st.radio("3. Apakah bermaksud melakukan pengujian rutin?", ["Ya", "Tidak"], index=["Ya", "Tidak"].index(st.session_state.q3), key="w_q3")
        st.write("*Jika Ya, sebutkan:*")
        c1, c2 = st.columns(2)
        with c1:
            st.text_input("Nama Instansi / Perusahaan:", value=st.session_state.q3_nama, key="w_q3_nama")
            st.text_input("No Telepon / HP:", value=st.session_state.q3_telp, key="w_q3_telp") 
        with c2:
            st.text_area("Alamat Lengkap:", value=st.session_state.q3_alamat, key="w_q3_alamat", height=105) 
            
        st.radio("4. Parameter memenuhi kebutuhan?", ["Cukup", "Kurang"], index=["Cukup", "Kurang"].index(st.session_state.q4), key="w_q4")
        st.write("*Jika Kurang, parameter apa yang perlu ditambahkan?*")
        c3, c4 = st.columns(2)
        for i, (param, baris) in enumerate(param_kurang.items()):
            if i % 2 == 0:
                with c3: st.checkbox(param, value=st.session_state[f"param_{baris}"], key=f"w_p_{baris}")
            else:
                with c4: st.checkbox(param, value=st.session_state[f"param_{baris}"], key=f"w_p_{baris}")
                
        st.text_input("Parameter Lain-lain:", value=st.session_state.q4_lainnya, key="w_q4_lainnya")
        st.write("---")
        st.text_area("5. Saran Peningkatan:", value=st.session_state.q5_saran, key="w_q5_saran")
        st.write("---")
        
        c_back, c_submit = st.columns([1, 3])
        with c_back:
            st.button("⬅️ Kembali", on_click=go_back)
        with c_submit:
            kirim = st.button("Kirim Kuesioner ✅", type="primary")

        if kirim:
            simpan_b() # Eksekusi simpan terakhir sebelum diproses
            try:
                # Cek nomor responden saat ini
                df_eval_count = len(pd.read_csv(DATA_F15))
                nomor_responden = df_eval_count + 1

                wb = openpyxl.load_workbook(TEMPLATE_F13)
                sheet = wb[NAMA_SHEET_F13]
                data_evaluasi_baru = {}
                
                # Memasukkan Nilai Bagian A (Kinerja)
                for idx, (baris, _, _) in enumerate(aspek_list):
                    harapan = st.session_state[f"h_{baris}"]
                    pelayanan = st.session_state[f"l_{baris}"]
                    sheet.cell(row=baris, column=9).value = harapan
                    sheet.cell(row=baris, column=11).value = pelayanan
                    data_evaluasi_baru[f'h_{idx+1}'] = harapan
                    data_evaluasi_baru[f'k_{idx+1}'] = pelayanan
                    
                nama_instansi = st.session_state.q3_nama if st.session_state.q3_nama else "Anonim"
                
                # Memasukkan Profil Pelanggan (Bagian B)
                if st.session_state.q1 == "Usaha": sheet.cell(row=50, column=3).value = "X"
                else: sheet.cell(row=51, column=3).value = "X"
                if st.session_state.q2 == "Laboratorium PDAM TKR": sheet.cell(row=54, column=3).value = "X"
                else: sheet.cell(row=55, column=3).value = "X"
                if st.session_state.q2_alasan: sheet.cell(row=56, column=5).value = st.session_state.q2_alasan
                if st.session_state.q3 == "Ya": 
                    sheet.cell(row=59, column=3).value = "X"
                    if st.session_state.q3_nama: sheet.cell(row=62, column=5).value = st.session_state.q3_nama
                    if st.session_state.q3_alamat: sheet.cell(row=63, column=5).value = st.session_state.q3_alamat 
                    if st.session_state.q3_telp: sheet.cell(row=64, column=5).value = st.session_state.q3_telp 
                else: sheet.cell(row=60, column=3).value = "X"
                    
                if st.session_state.q4 == "Cukup": sheet.cell(row=67, column=3).value = "X"
                else: 
                    sheet.cell(row=68, column=3).value = "X"
                    for param, baris in param_kurang.items():
                        if st.session_state[f"param_{baris}"]: sheet.cell(row=baris, column=3).value = "X"
                    if st.session_state.q4_lainnya: 
                        sheet.cell(row=80, column=3).value = "X"
                        sheet.cell(row=80, column=5).value = st.session_state.q4_lainnya

                if st.session_state.q5_saran: sheet.cell(row=82, column=3).value = st.session_state.q5_saran

                # Penamaan Otomatis: Responden_1, Responden_2, dst.
                nama_file_f13 = f"F13_Responden_{nomor_responden}_{nama_instansi}.xlsx"
                path_simpan = os.path.join(FOLDER_HASIL, nama_file_f13)
                wb.save(path_simpan)

                df_log = pd.read_csv(LOG_FILE)
                df_log = pd.concat([df_log, pd.DataFrame([{"Nama / Instansi": nama_instansi, "Tujuan Uji": st.session_state.q1}])], ignore_index=True)
                df_log.to_csv(LOG_FILE, index=False)

                df_eval = pd.read_csv(DATA_F15)
                df_eval = pd.concat([df_eval, pd.DataFrame([data_evaluasi_baru])], ignore_index=True)
                df_eval.to_csv(DATA_F15, index=False)

                st.success(f"Kuesioner Anda berhasil dikirim sebagai Responden {nomor_responden}! Terima kasih.")
                st.button("Isi Kuesioner Baru", on_click=reset_kuesioner)
                
            except Exception as e:
                st.error(f"Gagal memproses. Pastikan file ter-upload. Error: {e}")

# ==========================================
# HALAMAN 2: PANEL ADMIN
# ==========================================
elif menu == "Panel Admin":
    st.title("🛡️ Panel Admin")
    pwd = st.text_input("Masukkan Password:", type="password")
    
    if pwd == "admin123":
        st.success("Login Berhasil!")
        
        df_log = pd.read_csv(LOG_FILE)
        df_eval = pd.read_csv(DATA_F15)
        jumlah_responden = len(df_eval)
        st.write(f"### Total Responden Masuk: {jumlah_responden} Orang")
        
        if st.button("🖨️ Generate Evaluasi F15 (.xlsx)", type="primary"):
            if jumlah_responden > 0 and os.path.exists(TEMPLATE_F15):
                try:
                    baris_mulai_k = 12
                    baris_mulai_h = 36
                    baris_mulai_pemetaan = 71
                    
                    updates = {} 
                    for index, row in df_eval.iterrows():
                        if index >= 15: break
                        updates[(baris_mulai_k + index, 2)] = index + 1
                        updates[(baris_mulai_h + index, 2)] = index + 1
                        for i in range(1, 15):
                            updates[(baris_mulai_k + index, 2 + i)] = row[f'k_{i}']
                            updates[(baris_mulai_h + index, 2 + i)] = row[f'h_{i}']
                            
                    rata_rata_k = df_eval[[f'k_{i}' for i in range(1, 15)]].mean().tolist()
                    rata_rata_h = df_eval[[f'h_{i}' for i in range(1, 15)]].mean().tolist()
                    
                    for idx in range(14):
                        updates[(baris_mulai_pemetaan + idx, 4)] = rata_rata_k[idx]
                        updates[(baris_mulai_pemetaan + idx, 6)] = rata_rata_h[idx]
                    
                    wb_f15 = openpyxl.load_workbook(TEMPLATE_F15)
                    ws15 = wb_f15[NAMA_SHEET_F15]
                    
                    for (r, c), val in updates.items():
                        ws15.cell(row=r, column=c).value = val
                    
                    output_f15 = BytesIO()
                    wb_f15.save(output_f15)
                    output_f15.seek(0)
                    
                    st.success("File Evaluasi F15 berhasil dicetak!")
                    st.download_button(
                        label="📥 Download Hasil Evaluasi F15",
                        data=output_f15,
                        file_name=f"F15_Evaluasi_Lab_Total_{jumlah_responden}_Responden.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.info("Penting: Buka file Excel F15 ini, lalu copy-paste data tabelnya ke Template asli yang ada di komputer Anda agar grafiknya dapat terbaca sempurna.")
                except Exception as e:
                    st.error(f"Gagal mencetak F15: {e}")
            elif not os.path.exists(TEMPLATE_F15):
                st.error(f"File {TEMPLATE_F15} tidak ditemukan di server!")
            else:
                st.warning("Belum ada data kuesioner yang bisa dievaluasi.")
        
        st.write("---")
        st.write("### 📥 Download Data Mentah (F13)")
        if len(os.listdir(FOLDER_HASIL)) > 0:
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                for file_name in os.listdir(FOLDER_HASIL):
                    zip_file.write(os.path.join(FOLDER_HASIL, file_name), arcname=file_name)
            
            st.download_button("📦 Download Kumpulan F13 (.ZIP)", data=zip_buffer.getvalue(), file_name="Data_F13_Responden_Lengkap.zip")